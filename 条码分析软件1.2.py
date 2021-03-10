f = open('请英文格式下导入条码.txt')
lines = len(f.readlines())
f.seek(0)

DATE = []
SPMC = []
CPGG = []
GG2 = []
CS = []
TM = []

l1 = f.readlines()

a = 0
b = len(l1)
while a < b:
    l = l1[a].split('/')

    day = l[0][2:]
    ph = l[1]
    kj = l[2]
    dj = l[3]

    DATE.append(day)

    if dj == 'A':
        spmc = '一等品' + kj + '口径'
    elif dj == 'B':
        spmc = '二等品' + kj + '口径'
    else:
        spmc = '三等品' + kj + '口径'
    SPMC.append(spmc)

    m = len(l)
    if m == 6:
        dh = l[5][:12]
        gg2 = l[4]
        tm = day + '/' + ph + '/' + kj + '/' + dj + '/' + gg2 + '/' + dh
        GG2.append(gg2)
        TM.append(tm)
        CS.append(dh)
    
        if dj == 'A' and dh == '0772-6511119':
            dj1 = 'A'
        elif dj == 'A' and dh == '0772-6511099':
            dj1 = 'A1'
        elif dj == 'A' and dh == '0772-6511077':
            dj1 = 'A2'
        else:
            dj1 = dj
        cpgg = ph + dj1 
        CPGG.append(cpgg)
    elif m == 5:
        dh = '0'
        gg2 = l[4][:4]
        GG2.append(gg2)
        CS.append(dh)
        tm = day + '/' + ph + '/' + kj + '/' + dj + '/' + gg2
        TM.append(tm)
        cpgg = ph + dj
        CPGG.append(cpgg)
    
    a += 1
    
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Sheet1", 0)

ws1['A1'] = '商品编号'
ws1['B1'] = '商品名称'
ws1['C1'] = '单位'
ws1['D1'] = '产品规格'
ws1['E1'] = '规格2'
ws1['F1'] = '保质期'
ws1['G1'] = '商品条码'
ws1['H1'] = '最低库存'
ws1['I1'] = '预设进价'
ws1['J1'] = '预设售价'
ws1['K1'] = '一级价格'
ws1['L1'] = '二级价格'
ws1['M1'] = '三级价格'
ws1['N1'] = '四级价格'
ws1['O1'] = '五级价格'
ws1['P1'] = '生产厂商'
ws1['Q1'] = '备注'

v = 0

while v < lines:
	h = v + 2
	ws1.cell(row = h , column = 2).value = SPMC[v]
	ws1.cell(row = h , column = 3).value = '件'
	ws1.cell(row = h , column = 4).value = CPGG[v]
	ws1.cell(row = h , column = 5).value = GG2[v]
	ws1.cell(row = h , column = 6).value = 720
	ws1.cell(row = h , column = 7).value = TM[v]
	ws1.cell(row = h , column = 16).value = CS[v]
	ws1.cell(row = h , column = 17).value = DATE[v]
	
	v += 1

f.close()
wb.save('商品信息.xlsx')  