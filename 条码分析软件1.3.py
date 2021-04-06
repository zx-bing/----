import os
from openpyxl import Workbook

filepath = __file__
realpath = os.path.realpath(filepath)
current_path = os.path.dirname(realpath)

w = Workbook()
w.active
ws1 = w["Sheet"]
ws1.title = "商品信息模板"
ws1['A1'] = '商品编号'
ws1['B1'] = '商品名称'
ws1['C1'] = '单位'
ws1['D1'] = '产品规格'
ws1['E1'] = '规格2'
ws1['F1'] = '保质期'
ws1['G1'] = '商品条码'
ws1['H1'] = '最低库存'
ws1['I1'] = '预设进价'
ws1['J1'] = '预设售价'
ws1['K1'] = '零售价'
ws1['L1'] = '一级价格'
ws1['M1'] = '二级价格'
ws1['N1'] = '三级价格'
ws1['O1'] = '四级价格'
ws1['P1'] = '五级价格'
ws1['Q1'] = '生产厂商'
ws1['R1'] = '备注'
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['G'].width = 50
ws1.column_dimensions['Q'].width = 16
ws1.column_dimensions['R'].width = 16

f = open(current_path + '\\' + '请英文格式下导入条码.txt','r')
s = f.readlines()
f.close()
n = 1
for i in s:
    ws1.cell(row=n + 1, column=7).value = i
    ws1.cell(row=n + 1, column=3).value = '件'
    ws1.cell(row=n + 1, column=6).value = 720
    l = i.split('/')
    ws1.cell(row=n + 1, column=5).value = l[4][:4]
    ws1.cell(row=n + 1, column=18).value = l[0]
    if l[3] == 'A':
        spmc = '一等品' + l[2] + '口径'
    elif l[3] == 'B':
        spmc = '二等品' + l[2] + '口径'
    else:
        spmc = '三等品' + l[2] + '口径'
    ws1.cell(row=n + 1, column=2).value = spmc
    m = len(l)
    if m == 6:
        dh = l[5][:12]
    elif m == 5:
        dh = 0
    ws1.cell(row=n + 1, column=17).value = dh
    if l[3] == 'A' and dh == '0772-6511119':
        dj1 = 'A'
    elif l[3] == 'A' and dh == '0772-6511099':
        dj1 = 'A1'
    elif l[3] == 'A' and dh == '0772-6511077':
        dj1 = 'A2'
    else:
        dj1 = l[3]
    cpgg = l[1] + dj1
    ws1.cell(row=n + 1, column=4).value = cpgg
    n += 1
    print('正在解析条码...')
    print(i,end='')
    print(l)
    print('商品名称：{}'.format(spmc))
    print('商品单位：件')
    print('商品规格：{}'.format(cpgg))
    print('规格2：{}'.format(l[4][:4]))
    print('保质期：{}'.format(720))
    print('商品条码：{}'.format(i),end='')
    print('生产厂商：{}'.format(dh))
    print('备注：{}'.format(l[0]),end='\n\n\n')
import datetime

T = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
r = T[0:4] + T[5:7] + T[8:10] + T[11:13] + T[14:16] + T[17:]
print('写入 商品信息{}.xlsx 成功！'.format(r))
print('保存 商品信息{}.xlsx 成功！'.format(r))
w.save('商品信息{}.xlsx'.format(r))
os.system('pause')
